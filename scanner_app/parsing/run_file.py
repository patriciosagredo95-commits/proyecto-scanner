"""Parseo del contenido de un archivo RUN_XXXX.xlsx: metadata (lado izquierdo)
y tabla de productos (lado derecho, "Productos")."""

from dataclasses import dataclass
from datetime import datetime

import openpyxl
import pandas as pd

from scanner_app.config import (
    RUN_META_COL_ETIQUETA,
    RUN_META_COL_VALOR,
    RUN_META_FILAS,
    RUN_PRODUCTOS_COL_INICIO,
    RUN_PRODUCTOS_FILA_HEADER,
    RUN_PRODUCTOS_FILA_INICIO_DATOS,
    RUN_PRODUCTOS_HEADERS,
)


@dataclass(frozen=True)
class MetadataRun:
    run_numero_interno: int
    estado: str | None
    operador: str | None
    turno_texto: str | None
    comienzo: datetime | None
    fin: datetime | None


@dataclass(frozen=True)
class ArchivoRunParseado:
    hoja: str
    metadata: MetadataRun
    productos: pd.DataFrame  # todas las filas leídas (incluidas y excluidas)


def _normalizar_header(valor) -> str:
    return str(valor).strip() if valor is not None else ""


def _leer_valor_meta(ws, campo: str) -> str | None:
    fila_esperada, etiqueta_esperada = RUN_META_FILAS[campo]
    etiqueta_en_fila = ws.cell(row=fila_esperada, column=RUN_META_COL_ETIQUETA).value
    if etiqueta_en_fila is not None and str(etiqueta_en_fila).strip() == etiqueta_esperada:
        return ws.cell(row=fila_esperada, column=RUN_META_COL_VALOR).value

    # Fallback: la posición fija no coincidió (layout pudo haberse corrido);
    # escanear toda la columna A buscando la etiqueta exacta.
    for fila in range(1, ws.max_row + 1):
        valor = ws.cell(row=fila, column=RUN_META_COL_ETIQUETA).value
        if valor is not None and str(valor).strip() == etiqueta_esperada:
            return ws.cell(row=fila, column=RUN_META_COL_VALOR).value

    raise ValueError(
        f"No se encontró la etiqueta '{etiqueta_esperada}' esperada para el campo "
        f"'{campo}' en el archivo (ni en la fila {fila_esperada} ni en el resto de la columna A)."
    )


def _parse_datetime(valor) -> datetime | None:
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor
    return datetime.strptime(str(valor).strip(), "%d-%m-%Y %H:%M")


def _leer_metadata(ws) -> MetadataRun:
    run_numero_texto = _leer_valor_meta(ws, "run_numero")
    if run_numero_texto is None:
        raise ValueError("No se pudo leer el número de RUN interno del archivo.")
    digitos = "".join(ch for ch in str(run_numero_texto) if ch.isdigit())
    if not digitos:
        raise ValueError(f"El valor de RUN interno '{run_numero_texto}' no contiene dígitos.")
    run_numero_interno = int(digitos)

    return MetadataRun(
        run_numero_interno=run_numero_interno,
        estado=_leer_valor_meta(ws, "estado"),
        operador=_leer_valor_meta(ws, "operador"),
        turno_texto=_leer_valor_meta(ws, "turno"),
        comienzo=_parse_datetime(_leer_valor_meta(ws, "comienzo")),
        fin=_parse_datetime(_leer_valor_meta(ws, "fin")),
    )


def _leer_productos(ws) -> pd.DataFrame:
    # El ORDEN de las columnas E:V varía entre archivos reales (129 de 354
    # archivos reales de "Run 2026/**" traen "Cantidad"/"Volumen [%]"/"Pateador"
    # en otra posición que el resto -- mismo set de 18 columnas, reordenadas).
    # Por eso se ubica cada columna por NOMBRE de encabezado, no por posición
    # fija; solo se exige que el SET de nombres encontrados coincida con el
    # esperado (protección real ante un cambio de formato del scanner).
    ancho_tabla = len(RUN_PRODUCTOS_HEADERS)
    headers_encontrados = [
        _normalizar_header(ws.cell(row=RUN_PRODUCTOS_FILA_HEADER, column=RUN_PRODUCTOS_COL_INICIO + i).value)
        for i in range(ancho_tabla)
    ]
    headers_esperados = [_normalizar_header(h) for h in RUN_PRODUCTOS_HEADERS]
    if set(headers_encontrados) != set(headers_esperados):
        raise ValueError(
            "La tabla 'Productos' no tiene los encabezados esperados. "
            f"Encontrado: {headers_encontrados!r}. Esperado: {headers_esperados!r}. "
            "Es probable que el formato exportado por el scanner haya cambiado."
        )

    posicion_por_header_normalizado = {
        header_encontrado: RUN_PRODUCTOS_COL_INICIO + i for i, header_encontrado in enumerate(headers_encontrados)
    }
    col_index = {
        header: posicion_por_header_normalizado[_normalizar_header(header)] for header in RUN_PRODUCTOS_HEADERS
    }
    col_nombre = col_index["Nombre"]
    col_estado = col_index["Estado"]
    col_calidad = col_index["Calidad"]
    col_volumen_nominal_m3 = col_index["Volumen Nominal\n[ m³ ] "]
    col_cantidad_pcs = col_index["Cantidad\n[ pcs ] "]
    col_largo_pct = col_index["Largo\n[ % ] "]
    col_largo_m = col_index["Largo\n[ m ] "]
    col_largo_maximo = col_index["Largo Máximo"]
    col_largo_minimo = col_index["Largo Mínimo"]
    col_largo_promedio_m = col_index["Largo Promedio\n[ m ] "]
    col_volumen_nominal_pct = col_index["Volumen Nominal\n[ % ] "]
    col_volumen_m3 = col_index["Volumen\n[ m³ ] "]  # solo para el agregado a nivel de run, no se persiste por producto

    filas = []
    fila = RUN_PRODUCTOS_FILA_INICIO_DATOS
    while True:
        nombre = ws.cell(row=fila, column=col_nombre).value
        if nombre is None or str(nombre).strip() == "":
            break
        cantidad_pcs = ws.cell(row=fila, column=col_cantidad_pcs).value or 0
        volumen_nominal_m3 = ws.cell(row=fila, column=col_volumen_nominal_m3).value
        estado = ws.cell(row=fila, column=col_estado).value
        # Regla de inclusión: solo Cantidad y Volumen Nominal > 0 -- el campo
        # Estado NO se usa para filtrar (confirmado contra RUN_2830--EJEMPLO,
        # que incluye una fila "Inactivo" con cantidad y volumen > 0).
        incluido = cantidad_pcs > 0 and (volumen_nominal_m3 or 0) > 0
        filas.append(
            {
                "nombre": str(nombre).strip(),
                "estado": estado,
                "calidad": ws.cell(row=fila, column=col_calidad).value,
                "volumen_nominal_m3": volumen_nominal_m3,
                "cantidad_pcs": cantidad_pcs,
                "largo_pct": ws.cell(row=fila, column=col_largo_pct).value,
                "largo_m": ws.cell(row=fila, column=col_largo_m).value,
                "largo_maximo": ws.cell(row=fila, column=col_largo_maximo).value,
                "largo_minimo": ws.cell(row=fila, column=col_largo_minimo).value,
                "largo_promedio_m": ws.cell(row=fila, column=col_largo_promedio_m).value,
                "volumen_nominal_pct": ws.cell(row=fila, column=col_volumen_nominal_pct).value,
                "volumen_m3": ws.cell(row=fila, column=col_volumen_m3).value,
                "incluido": incluido,
            }
        )
        fila += 1

    return pd.DataFrame(filas)


def parse_run_file(archivo) -> ArchivoRunParseado:
    """archivo: ruta o file-like (ej. UploadedFile de Streamlit)."""
    wb = openpyxl.load_workbook(archivo, data_only=True)
    hoja = wb.sheetnames[0]
    ws = wb[hoja]
    metadata = _leer_metadata(ws)
    productos = _leer_productos(ws)
    return ArchivoRunParseado(hoja=hoja, metadata=metadata, productos=productos)
